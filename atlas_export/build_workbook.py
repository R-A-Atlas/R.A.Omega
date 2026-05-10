"""Build Excel workbook from POST /query-shaped analysis JSON."""

from __future__ import annotations

import json
import re
from datetime import date
from pathlib import Path
from typing import Any

ROOT = Path(__file__).resolve().parent.parent
MODELS_DIR = ROOT / "atlas_vault" / "03-Outputs" / "Models"


def _tick(envelope: dict[str, Any]) -> str:
    pq = envelope.get("parsed_query") or {}
    if isinstance(pq, dict):
        t = pq.get("tickers")
        if isinstance(t, (list, tuple)) and t:
            return str(t[0]).upper().strip()[:12]
    fr = envelope.get("final_report") or {}
    if isinstance(fr, dict) and fr.get("ticker"):
        return str(fr["ticker"]).upper().strip()[:12]
    return "ATLAS"


def write_query_envelope_xlsx(
    envelope: dict[str, Any],
    dest: Path | None = None,
) -> Path:
    from openpyxl import Workbook
    from openpyxl.styles import Font

    MODELS_DIR.mkdir(parents=True, exist_ok=True)
    tk = re.sub(r"[^\w\-]+", "_", _tick(envelope))[:32] or "ATLAS"
    target = dest or (MODELS_DIR / f"{tk}_{date.today().strftime('%Y-%m-%d')}.xlsx")

    fr = envelope.get("final_report") if isinstance(envelope.get("final_report"), dict) else {}
    wb = Workbook()

    # Summary
    ws0 = wb.active
    ws0.title = "Summary"
    bold = Font(bold=True)
    ws0.append(["Field", "Value"])
    ws0["A1"].font = bold
    ws0["B1"].font = bold
    rows = [
        ("Query", str(envelope.get("query") or "")),
        ("Ticker", _tick(envelope)),
        ("TLDR", str(envelope.get("tldr") or fr.get("tldr") or "")),
        ("Overall rating", str(fr.get("overall_rating") or "")),
        ("Confidence", str(fr.get("confidence") or "")),
        ("Price now", str(fr.get("price_now") or "")),
    ]
    for r in rows:
        ws0.append(list(r))

    # Trade Plan
    ws1 = wb.create_sheet("Trade Plan")
    ws1.append(["Key", "Value"])
    ws1["A1"].font = bold
    ws1["B1"].font = bold
    tp = fr.get("trade_plan")
    if isinstance(tp, dict):
        for k, v in tp.items():
            ws1.append([str(k), str(v)])
    elif isinstance(tp, str) and tp.strip():
        ws1.append(["plan", tp.strip()])
    else:
        ws1.append(["trade_plan", json.dumps(tp)[:2000] if tp is not None else ""])

    # Scenario Analysis
    ws2 = wb.create_sheet("Scenario Analysis")
    ws2.append(["Label", "Probability", "Trigger", "Outcome", "Your action"])
    for c in ws2[1]:
        c.font = bold
    arr = envelope.get("scenarios")
    if isinstance(arr, list):
        for sc in arr[:20]:
            if not isinstance(sc, dict):
                ws2.append([str(sc), "", "", "", ""])
                continue
            ws2.append(
                [
                    str(sc.get("label") or sc.get("name") or ""),
                    str(sc.get("probability") or sc.get("prob") or ""),
                    str(sc.get("trigger") or ""),
                    str(sc.get("outcome") or ""),
                    str(sc.get("your_action") or ""),
                ]
            )

    # Price Levels
    ws3 = wb.create_sheet("Price Levels")
    ws3.append(["Level", "Value"])
    ws3["A1"].font = bold
    ws3["B1"].font = bold
    pl = fr.get("price_levels")
    if isinstance(pl, dict):
        for k, v in pl.items():
            ws3.append([str(k), str(v)])
    elif isinstance(pl, str) and pl.strip():
        ws3.append(["raw", pl.strip()])

    wb.save(str(target.resolve()))
    return target


__all__ = ["write_query_envelope_xlsx", "MODELS_DIR"]
